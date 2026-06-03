import os
import sys
import logging
import re
from PIL import Image, ImageTk

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


def _get_resource_dir():
    """获取 timeline_tool 资源目录。打包后从 exe 同级目录查找，开发时从源码目录查找。"""
    if getattr(sys, 'frozen', False):
        # PyInstaller 打包后：exe 所在目录即为资源根目录
        return os.path.dirname(sys.executable)
    else:
        # 源码运行：timeline_tool/ 子目录
        return os.path.dirname(os.path.abspath(__file__))


_RES_DIR = _get_resource_dir()
PORTRAIT_DIR = os.path.join(_RES_DIR, "operator_portraits")
ALIAS_FILE = os.path.join(_RES_DIR, "operator_aliases.xlsx")


class OperatorPortraitManager:
    """干员头像管理器：扫描头像目录、加载昵称映射、提供头像查找。"""

    def __init__(self, scaling_factor=1.0):
        self.scaling_factor = scaling_factor
        # 头像尺寸与字体高度相近，确保和其他文字比例协调
        self.portrait_size = max(20, int(22 * scaling_factor))

        self._id_to_path = {}      # 编号 -> 路径
        self._name_to_path = {}    # 名称 -> 路径
        self._alias_to_name = {}   # 昵称/别名 -> 标准名称
        self._cache = {}           # 路径 -> PhotoImage

        self._load_aliases()
        self._load_portraits()

    def _load_aliases(self):
        """从 Excel 加载昵称映射。格式：第一列为标准名称，后续列为该名称的多个别名。"""
        if not HAS_OPENPYXL:
            logger.warning("openpyxl 未安装，无法加载昵称映射。")
            return
        if not os.path.exists(ALIAS_FILE):
            logger.warning(f"昵称映射文件不存在: {ALIAS_FILE}")
            return
        try:
            wb = load_workbook(ALIAS_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                name = str(row[0]).strip()
                if not name:
                    continue
                # 同一行后续所有列都是该名称的别名
                for cell in row[1:]:
                    if cell is not None:
                        alias = str(cell).strip()
                        if alias:
                            self._alias_to_name[alias] = name
            logger.info(f"已加载 {len(self._alias_to_name)} 条昵称映射")
        except Exception as e:
            logger.error(f"加载昵称映射失败: {e}")

    def _load_portraits(self):
        """扫描 operator_portraits/ 目录，解析文件名 编号_名称.扩展名。"""
        if not os.path.exists(PORTRAIT_DIR):
            logger.warning(f"头像目录不存在: {PORTRAIT_DIR}，请手动放入干员头像图片")
            return
        for fname in os.listdir(PORTRAIT_DIR):
            if not fname.lower().endswith(('.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp')):
                continue
            path = os.path.join(PORTRAIT_DIR, fname)
            base, _ = os.path.splitext(fname)
            parts = base.split('_', 1)
            if len(parts) == 2:
                op_id, op_name = parts[0].strip(), parts[1].strip()
                if op_id:
                    self._id_to_path[op_id] = path
                if op_name:
                    self._name_to_path[op_name] = path
            elif len(parts) == 1:
                self._name_to_path[parts[0]] = path
        logger.info(
            f"已扫描 {len(self._id_to_path)} 个编号头像, {len(self._name_to_path)} 个名称头像"
        )

    def resolve(self, key):
        """根据编号/名称/昵称查找头像文件路径，找不到返回 None。"""
        key = str(key).strip()
        # 直接匹配编号
        if key in self._id_to_path:
            return self._id_to_path[key]
        # 直接匹配名称
        if key in self._name_to_path:
            return self._name_to_path[key]
        # 查别名表
        real_name = self._alias_to_name.get(key)
        if real_name and real_name in self._name_to_path:
            return self._name_to_path[real_name]
        return None

    def get_image(self, path):
        """获取指定路径的 PhotoImage（带缓存和缩放）。"""
        if path in self._cache:
            return self._cache[path]
        try:
            img = Image.open(path).convert("RGBA")
            img = img.resize((self.portrait_size, self.portrait_size), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self._cache[path] = photo
            return photo
        except Exception as e:
            logger.error(f"加载头像失败 {path}: {e}")
            return None

    def parse_name(self, raw_name):
        """解析节点名称中的 [...] 标记。

        Returns:
            (display_name, portrait_path)
            portrait_path 为 None 表示未匹配到头像。
        """
        if not raw_name:
            return raw_name, None
        match = re.search(r'\[(.*?)\]', raw_name)
        if not match:
            return raw_name, None
        key = match.group(1).strip()
        path = self.resolve(key)
        if not path:
            return raw_name, None
        display_name = raw_name.replace(f'[{key}]', '').strip()
        return display_name, path
