#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 轴文件转 JSON 工具
用法: 直接运行，弹出对话框选择 Excel 文件，自动在同目录生成同名 .json
"""

import json
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox

import openpyxl

# 把项目根目录加入路径，以便导入 timeline_tool 的模块
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
if os.path.basename(SCRIPT_DIR) == "小工具":
    sys.path.insert(0, PROJECT_ROOT)

try:
    from timeline_tool.config import FPS
except ImportError:
    FPS = 30


def parse_frame_time(time_str):
    """将 MM:SS:FF 格式解析为总帧数，失败则尝试直接转为 int。"""
    if isinstance(time_str, int):
        return time_str
    if isinstance(time_str, str):
        time_str = time_str.strip()
        parts = time_str.split(":")
        if len(parts) == 3:
            try:
                minutes = int(parts[0])
                seconds = int(parts[1])
                frames = int(parts[2])
                return minutes * 60 * FPS + seconds * FPS + frames
            except ValueError:
                pass
        # 尝试直接解析为整数帧
        try:
            return int(time_str)
        except ValueError:
            pass
    return 0


def excel_to_json(excel_path, json_path):
    """将 Excel 轴文件转换回 JSON 文件。"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ========== 读取设置 Sheet ==========
    ws_settings = wb["设置"] if "设置" in wb.sheetnames else None
    version = "2.0"
    track_settings = []

    if ws_settings:
        # 第一行: 版本信息
        for row in ws_settings.iter_rows(min_row=1, max_row=1, values_only=True):
            if row and len(row) >= 2 and str(row[0]).strip() == "版本":
                version = str(row[1]).strip() if row[1] else "2.0"

        # 第三行是表头，从第四行开始是数据
        for row in ws_settings.iter_rows(min_row=4, values_only=True):
            if not row or not row[0]:
                continue
            track_settings.append({
                "name": str(row[0]).strip() if row[0] else "默认轨道",
                "mode": str(row[1]).strip() if len(row) > 1 and row[1] else "打轴模式",
                "magnet_mode": str(row[2]).strip() in ("开", "True", "true", "1") if len(row) > 2 and row[2] else True,
                "sound_alert_enabled": str(row[3]).strip() in ("开", "True", "true", "1") if len(row) > 3 and row[3] else True,
                "visual_alert_enabled": str(row[4]).strip() in ("开", "True", "true", "1") if len(row) > 4 and row[4] else True,
                "alert_lead_frames": int(row[5]) if len(row) > 5 and row[5] is not None else 60,
            })

    # ========== 读取时间轴 Sheet ==========
    ws_timeline = wb["时间轴"] if "时间轴" in wb.sheetnames else wb.active
    rows = list(ws_timeline.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 文件为空或没有数据行")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    # 期望 headers: ["时间(帧)", "时间(MM:SS:FF)", 轨道0名称, 轨道1名称, ...]
    if len(headers) < 3:
        raise ValueError("时间轴表头格式不正确，至少需要 3 列")

    track_names = headers[2:]
    num_tracks = len(track_names)

    # 如果没有设置 Sheet，则根据表头生成默认设置
    if not track_settings:
        track_settings = [
            {"name": name, "mode": "打轴模式", "magnet_mode": True,
             "sound_alert_enabled": True, "visual_alert_enabled": True, "alert_lead_frames": 60}
            for name in track_names
        ]

    # 初始化各轨道节点列表
    tracks_nodes = [[] for _ in range(num_tracks)]

    # 颜色轮询（与 config.NODE_COLORS 一致）
    NODE_COLORS = ["#ff6347", "#4682b4", "#32cd32", "#ffd700", "#9370db", "#ffa500"]

    for data_row in rows[1:]:
        if not data_row or data_row[0] is None:
            continue
        frame = parse_frame_time(data_row[0])
        for t_idx in range(num_tracks):
            cell_idx = 2 + t_idx
            if cell_idx < len(data_row) and data_row[cell_idx]:
                node_name = str(data_row[cell_idx]).strip()
                if node_name:
                    color = NODE_COLORS[len(tracks_nodes[t_idx]) % len(NODE_COLORS)]
                    tracks_nodes[t_idx].append({
                        "frame": frame,
                        "name": node_name,
                        "color": color
                    })

    # 组装 JSON
    tracks = []
    for t_idx in range(num_tracks):
        settings = track_settings[t_idx] if t_idx < len(track_settings) else {}
        tracks.append({
            "name": settings.get("name", track_names[t_idx] if t_idx < len(track_names) else f"轨道{t_idx}"),
            "mode": settings.get("mode", "打轴模式"),
            "magnet_mode": settings.get("magnet_mode", True),
            "sound_alert_enabled": settings.get("sound_alert_enabled", True),
            "visual_alert_enabled": settings.get("visual_alert_enabled", True),
            "alert_lead_frames": settings.get("alert_lead_frames", 60),
            "nodes": tracks_nodes[t_idx]
        })

    payload = {
        "version": version,
        "tracks": tracks
    }

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, indent=4, ensure_ascii=False)

    return True


def main():
    root = tk.Tk()
    root.withdraw()
    root.wm_attributes("-topmost", True)

    excel_path = filedialog.askopenfilename(
        title="选择 Excel 轴文件",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        parent=root
    )
    if not excel_path:
        return

    base, _ = os.path.splitext(excel_path)
    json_path = base + ".json"

    counter = 1
    original_path = json_path
    while os.path.exists(json_path):
        json_path = f"{base}_{counter}.json"
        counter += 1

    try:
        excel_to_json(excel_path, json_path)
        messagebox.showinfo("转换成功", f"已保存到:\n{json_path}", parent=root)
    except Exception as e:
        messagebox.showerror("转换失败", f"错误信息:\n{e}", parent=root)
        raise


if __name__ == "__main__":
    main()
