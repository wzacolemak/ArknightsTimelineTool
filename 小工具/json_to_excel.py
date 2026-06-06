#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JSON 轴文件转 Excel 工具
用法: 直接运行，弹出对话框选择 JSON 文件，自动在同目录生成同名 .xlsx
"""

import json
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# 把项目根目录加入路径，以便导入 timeline_tool 的模块
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
if os.path.basename(SCRIPT_DIR) == "小工具":
    sys.path.insert(0, PROJECT_ROOT)

try:
    from timeline_tool.utils import format_frame_time
except ImportError:
    # 备用：自行定义
    FPS = 30
    def format_frame_time(total_frames):
        if not isinstance(total_frames, int) or total_frames < 0:
            return "--:--:--"
        total_seconds = total_frames // FPS
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        frames = total_frames % FPS
        return f"{minutes:02d}:{seconds:02d}:{frames:02d}"


def json_to_excel(json_path, excel_path):
    """将 JSON 轴文件转换为 Excel 文件。"""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 兼容旧版：纯列表视为单轨道
    if isinstance(data, list):
        tracks = [{"name": "默认轨道", "nodes": data}]
        version = "1.0"
    elif isinstance(data, dict) and "tracks" in data:
        tracks = data["tracks"]
        version = data.get("version", "2.0")
    else:
        tracks = [{"name": "默认轨道", "nodes": data if isinstance(data, list) else []}]
        version = "1.0"

    wb = openpyxl.Workbook()

    # ========== Sheet 1: 时间轴 ==========
    ws1 = wb.active
    ws1.title = "时间轴"

    # 表头
    headers = ["时间(帧)", "时间(MM:SS:FF)"]
    for track in tracks:
        headers.append(track.get("name", "未命名轨道"))
    ws1.append(headers)

    # 表头样式
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 收集所有出现过的帧数
    all_frames = set()
    for track in tracks:
        for node in track.get("nodes", []):
            all_frames.add(node.get("frame", 0))

    sorted_frames = sorted(all_frames)

    # 建立 frame -> (track_idx -> node_name) 映射
    frame_to_nodes = {}
    for frame in sorted_frames:
        frame_to_nodes[frame] = {}

    for t_idx, track in enumerate(tracks):
        for node in track.get("nodes", []):
            frame = node.get("frame", 0)
            name = node.get("name", "")
            # 同一轨道同一帧理论上只有一个节点，如有多个保留最后一个
            frame_to_nodes[frame][t_idx] = name

    # 填充数据行
    for frame in sorted_frames:
        row = [frame, format_frame_time(frame)]
        for t_idx in range(len(tracks)):
            row.append(frame_to_nodes[frame].get(t_idx, ""))
        ws1.append(row)

    # 自动调整列宽
    for col in ws1.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws1.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # ========== Sheet 2: 设置 ==========
    ws2 = wb.create_sheet(title="设置")
    ws2.append(["版本", version])
    ws2.append([])
    ws2.append(["轨道名称", "模式", "磁铁模式", "声音提醒", "视觉提醒", "提前帧数"])

    # 表头样式
    for cell in ws2[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for track in tracks:
        mode = track.get("mode", "打轴模式")
        magnet = "开" if track.get("magnet_mode", True) else "关"
        sound = "开" if track.get("sound_alert_enabled", True) else "关"
        visual = "开" if track.get("visual_alert_enabled", True) else "关"
        lead = track.get("alert_lead_frames", 60)
        ws2.append([track.get("name", ""), mode, magnet, sound, visual, lead])

    # 自动调整列宽
    for col in ws2.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws2.column_dimensions[col_letter].width = min(max_length + 4, 30)

    wb.save(excel_path)
    return True


def main():
    root = tk.Tk()
    root.withdraw()
    root.wm_attributes("-topmost", True)

    json_path = filedialog.askopenfilename(
        title="选择 JSON 轴文件",
        filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        parent=root
    )
    if not json_path:
        return

    base, _ = os.path.splitext(json_path)
    excel_path = base + ".xlsx"

    # 如果已存在则加数字后缀
    counter = 1
    original_path = excel_path
    while os.path.exists(excel_path):
        excel_path = f"{base}_{counter}.xlsx"
        counter += 1

    try:
        json_to_excel(json_path, excel_path)
        messagebox.showinfo("转换成功", f"已保存到:\n{excel_path}", parent=root)
    except Exception as e:
        messagebox.showerror("转换失败", f"错误信息:\n{e}", parent=root)
        raise


if __name__ == "__main__":
    main()
